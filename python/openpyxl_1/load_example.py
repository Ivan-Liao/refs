from openpyxl import load_workbook

workbook = load_workbook(filename="fin_sample.xlsx")
print(workbook.sheetnames)

sheet = workbook.active 
print(sheet)
print(sheet.title)

print(sheet['A1'])
print(sheet['A1'].value)
print(sheet['F10'].value)

print(sheet.cell(row=10, column=6))
print(sheet.cell(row=10, column=6).value)